import pyautogui, tkinter, pyperclip, webbrowser, time
import openpyxl
import os

window = tkinter.Tk()
window.title("Stocks")
label = tkinter.Label(window, text = 'Stock Data Entry').pack()
window.mainloop()

"""
directory = input('Excel file directory: ')
#initialize excel file
os.chdir(directory)
wb = openpyxl.load_workbook(filename = 'test.xlsx')


#loop to fill init list
loop = True
tickList = []
while(loop):
    #--------CAPITALIZE INPUT
    ticker = input("Input the ticker(s) you want to search (input \"ZZZZ\" to end list):")
    if (ticker == "ZZZZ"):
        break;
    else:
        tickList.append(ticker.upper())

r = 1
for i in tickList:
    #-Beginning of loop with initial list
    webbrowser.open('https://www.finance.yahoo.com/quote/' + i)
    pyautogui.moveTo(546, 726)#--------------- 1080 x 1920
    time.sleep(5.2) #lets the webpage load

    pyautogui.click(clicks = 3)
    pyautogui.hotkey('ctrl', 'c') #takes the hi lo and puts into the clipboard

    hilow = pyperclip.paste()
    print(hilow)

    
    split = hilow.split("-")


    pyautogui.moveTo(498, 406)#------------------1080 x 1920
    pyautogui.click(clicks = 3)
    pyautogui.hotkey('ctrl', 'c')

    price = pyperclip.paste()
    s = price.find('.')
    print(s)
    xPrice = price[:(s+3)]
    fPrice = xPrice.replace(',','')
    print(fPrice)

    pyautogui.click(x = 430, y = 15) #close the tab

    ws = wb.active
    ws.cell(row = r, column = 1).value = str(tickList[r-1])
    ws.cell(row = r, column = 2).value = float(fPrice)

    
    ws.cell(row = r, column = 3).value = float(split[0].strip().replace(',',''))
    ws.cell(row = r, column = 4).value = float(split[1].strip().replace(',',''))
    wb.save('test.xlsx')
    
    r += 1
"""
